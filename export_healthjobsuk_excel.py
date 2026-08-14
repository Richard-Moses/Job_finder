"""Builds healthjobsuk_jobs.xlsx from healthjobsuk_merged.json."""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(SCRIPT_DIR, "healthjobsuk_merged.json"), encoding="utf-8") as fh:
    jobs = json.load(fh)

headers = [
    "Title",
    "Grade / Band",
    "Employer",
    "Location",
    "Speciality",
    "Salary",
    "Sponsorship",
    "Sponsorship Snippet",
    "Job URL",
]
field_order = [
    "title",
    "grade",
    "employer",
    "location",
    "speciality",
    "salary",
    "sponsorship",
    "sponsorship_snippet",
    "jobDetailUrl",
]

wb = Workbook()
ws = wb.active
ws.title = "HealthJobsUK Jobs"

ws.append(headers)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"

for job in jobs:
    ws.append([job.get(f, "") for f in field_order])

ws.sheet_format.defaultRowHeight = 18
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 18
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).alignment = Alignment(wrap_text=False, vertical="center")

url_col = field_order.index("jobDetailUrl") + 1
for r in range(2, ws.max_row + 1):
    cell = ws.cell(row=r, column=url_col)
    if cell.value:
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

widths = {
    "Title": 45,
    "Grade / Band": 18,
    "Employer": 42,
    "Location": 24,
    "Speciality": 40,
    "Salary": 38,
    "Sponsorship": 20,
    "Sponsorship Snippet": 60,
    "Job URL": 50,
}
for i, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 20)

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

out_path = os.path.join(SCRIPT_DIR, "healthjobsuk_jobs.xlsx")
wb.save(out_path)
print(f"Wrote {len(jobs)} jobs to {out_path}")
