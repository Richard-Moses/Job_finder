"""
Scrapes job listings from an arbitrary URL via Firecrawl's structured
extraction and writes them to an Excel file.

Unlike scrape_dailyremote.py (which is hardcoded to that site's exact HTML)
this uses Firecrawl's LLM-based extraction, which adapts to different page
layouts -- so it works across many job board designs without custom parsing
code. It is NOT magic, though:

- Field coverage varies by site: some sites show salary/company on the
  listing page, some hide it behind a detail page (this script only reads
  the page(s) you give it, not per-job detail pages).
- Pagination is not auto-detected. Every site does it differently
  (?page=2, &_pg=2, /page/2/, infinite scroll, a "load more" button that
  needs JavaScript...). You tell this script the query-parameter name your
  target site uses (e.g. "page", "_pg", "p") and it appends
  &<param>=<N> for pages 2, 3, .... If a site doesn't use that pattern
  (or uses infinite scroll), pagination beyond page 1 will silently return
  duplicate or empty results -- check page 2's output before scaling up
  the page count.

Usage:
    python scrape_custom.py --url "https://example.com/jobs" --pages 1
    python scrape_custom.py --url "https://example.com/jobs?search=nurse" --pages 5 --page-param page
"""

import argparse
import csv
import json
import os
import sys
import time
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def load_dotenv(path):
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip())


load_dotenv(os.path.join(SCRIPT_DIR, ".env"))

FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY")
if not FIRECRAWL_API_KEY:
    sys.exit("Set FIRECRAWL_API_KEY in JOBFINDER/.env or the environment before running this script.")

FIRECRAWL_SCRAPE_URL = "https://api.firecrawl.dev/v1/scrape"
DELAY = 0.5
RETRIES = 3
FIELDNAMES = ["title", "company", "location", "job_type", "salary", "url"]

EXTRACTION_PROMPT = (
    "Extract every individual job listing shown on this page (skip navigation, "
    "footer, and any 'featured'/'sponsored' promo blocks that aren't real listings "
    "unless they are). For each job return: title, company (the hiring "
    "organisation/employer name, empty string if not shown), location, job_type "
    "(e.g. Full-Time/Part-Time/Contract, empty string if not shown), salary "
    "(empty string if not shown), and url (the absolute URL of the job's own "
    "detail page)."
)


def with_page_param(url, param, page):
    if page <= 1:
        return url
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query))
    query[param] = str(page)
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


def firecrawl_json_scrape(url):
    payload = {
        "url": url,
        "formats": ["json"],
        "onlyMainContent": True,
        "proxy": "stealth",
        "jsonOptions": {"prompt": EXTRACTION_PROMPT},
    }
    headers = {
        "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
        "Content-Type": "application/json",
    }
    for attempt in range(1, RETRIES + 1):
        try:
            resp = requests.post(FIRECRAWL_SCRAPE_URL, json=payload, headers=headers, timeout=60)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("success"):
                    j = data.get("data", {}).get("json", {})
                    for key in ("jobListings", "jobs", "job_listings", "listings"):
                        if key in j and isinstance(j[key], list):
                            return j[key]
                    for v in j.values():
                        if isinstance(v, list):
                            return v
                    return []
                print(f"  [warn] firecrawl error: {data}", file=sys.stderr)
            else:
                print(f"  [warn] HTTP {resp.status_code}: {resp.text[:200]}", file=sys.stderr)
        except requests.RequestException as exc:
            print(f"  [warn] {exc}", file=sys.stderr)
        time.sleep(2 * attempt)
    return None


def export_to_excel(jobs, out_path):
    headers = ["Title", "Company", "Location", "Job Type", "Salary", "Job URL"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Custom Jobs"
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    for job in jobs:
        ws.append([job.get(f, "") for f in FIELDNAMES[:-1]] + [job.get("url", "")])

    ws.sheet_format.defaultRowHeight = 18
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=False, vertical="center")
        url_cell = ws.cell(row=r, column=len(headers))
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"

    widths = {"Title": 45, "Company": 32, "Location": 24, "Job Type": 16, "Salary": 28, "Job URL": 55}
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 20)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--url", required=True, help="the job listings page URL")
    parser.add_argument("--pages", type=int, default=1, help="how many pages to fetch")
    parser.add_argument("--page-param", default="page", help="query param used for pagination, e.g. page, _pg, p")
    parser.add_argument("--out", default=os.path.join(SCRIPT_DIR, "custom_jobs.xlsx"))
    args = parser.parse_args()

    all_jobs = []
    seen = set()
    for page in range(1, args.pages + 1):
        page_url = with_page_param(args.url, args.page_param, page)
        print(f"[page {page}/{args.pages}] {page_url}")
        jobs = firecrawl_json_scrape(page_url)
        if jobs is None:
            print("  [error] fetch failed, stopping")
            break
        if not jobs:
            print("  [info] no jobs found on this page, stopping")
            break
        new_count = 0
        for job in jobs:
            key = job.get("url") or json.dumps(job, sort_keys=True)
            if key in seen:
                continue
            seen.add(key)
            all_jobs.append({k: job.get(k, "") for k in FIELDNAMES})
            new_count += 1
        print(f"  [info] +{new_count} new jobs (total {len(all_jobs)})")
        if page == 2 and new_count == 0:
            print(
                "  [warn] page 2 returned nothing new -- this site's pagination "
                "may not use the query param you specified. Stopping early."
            )
            break
        time.sleep(DELAY)

    print(f"== Writing {len(all_jobs)} jobs to {args.out} ==")
    export_to_excel(all_jobs, args.out)
    print("== Done ==")


if __name__ == "__main__":
    main()
