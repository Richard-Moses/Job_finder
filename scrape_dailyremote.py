"""
Scrapes remote job listings from dailyremote.com and exports them to an Excel
workbook.

Usage:
    python scrape_dailyremote.py [--target N] [--search "keyword"]

Notes:
- Listing pages (dailyremote.com/remote-jobs?search=...&page=N) expose almost
  every field for free (title, location, job type, salary, experience level,
  category, skill tags, a short AI summary) but hide the company name behind
  a "premium" paywall.
- Each job's own detail page (dailyremote.com/remote-job/<slug>) reveals the
  real company name and the full job description for free (verified via a
  plain, logged-out HTTP request), via an embedded JSON-LD <script> block.
- So this script does two passes: crawl listing pages to build the job list,
  then visit each job's detail page once to fill in company + full
  description.
- robots.txt for this site only disallows /apply/ (which we never touch);
  /remote-jobs and /remote-job/ are allowed.
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from html import unescape

import requests
from bs4 import BeautifulSoup

# Windows consoles/redirected files often default to cp1252, which chokes on
# non-ASCII job titles (é, ü, etc.) unless we force UTF-8 output.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

BASE = "https://dailyremote.com"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
}
DELAY = 0.4  # seconds between requests, be polite to the server
RETRIES = 3
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHECKPOINT_CSV = os.path.join(SCRIPT_DIR, "jobs_checkpoint.csv")

FIELDNAMES = [
    "job_id",
    "title",
    "company",
    "location",
    "job_type",
    "salary",
    "experience_level",
    "category",
    "role_tag",
    "skill_tags",
    "date_posted",
    "posted_relative",
    "application_deadline",
    "summary",
    "description",
    "url",
]

session = requests.Session()
session.headers.update(HEADERS)


def fetch(url):
    for attempt in range(1, RETRIES + 1):
        try:
            resp = session.get(url, timeout=20)
            if resp.status_code == 200:
                return resp.text
            print(f"  [warn] HTTP {resp.status_code} for {url}", file=sys.stderr)
        except requests.RequestException as exc:
            print(f"  [warn] {exc} for {url}", file=sys.stderr)
        time.sleep(1.5 * attempt)
    return None


def parse_listing_page(html):
    soup = BeautifulSoup(html, "html.parser")
    jobs = []
    for art in soup.select("article.card.js-card"):
        job_id = (art.get("data-id") or "").strip()
        if not job_id:
            continue

        h2_a = art.select_one("h2.job-position a")
        title = h2_a.get_text(strip=True) if h2_a else ""
        href = h2_a["href"] if h2_a and h2_a.has_attr("href") else ""
        url = BASE + href if href.startswith("/") else href

        job_type = ""
        posted_relative = ""
        company_name_div = art.select_one("div.company-name")
        if company_name_div:
            for sp in company_name_div.find_all("span"):
                style = sp.get("style", "")
                txt = sp.get_text(strip=True)
                if not txt or txt == "·":
                    continue
                if "margin-left" in style:
                    posted_relative = txt
                elif "margin-right" in style:
                    job_type = txt

        location = salary = experience = ""
        for tag in art.select("span.card-tag"):
            txt = tag.get_text(" ", strip=True)
            if txt.startswith("\U0001f30e"):  # 🌎
                location = txt.replace("\U0001f30e", "").strip()
            elif txt.startswith("\U0001f4b5"):  # 💵
                salary = txt.replace("\U0001f4b5", "").strip()
            elif txt.startswith("⭐"):  # ⭐
                experience = txt.replace("⭐", "").strip()

        cat_a = art.select_one("span.category-tag a")
        category = ""
        if cat_a:
            category = cat_a.get_text(strip=True).replace("\U0001f4bc", "").strip()

        role_a = art.select_one("a.role-tag")
        role_tag = role_a.get_text(strip=True) if role_a else ""

        summary_div = art.select_one("div.ai-responsibilities")
        summary = summary_div.get_text(strip=True) if summary_div else ""

        skills = [
            a.get_text(strip=True)
            for a in art.select("div.tags-container a")
            if "tags" in (a.get("class") or [])
        ]

        jobs.append(
            {
                "job_id": job_id,
                "title": title,
                "company": "",
                "location": location,
                "job_type": job_type,
                "salary": salary,
                "experience_level": experience,
                "category": category,
                "role_tag": role_tag,
                "skill_tags": ", ".join(skills),
                "date_posted": "",
                "posted_relative": posted_relative,
                "application_deadline": "",
                "summary": summary,
                "description": "",
                "url": url,
            }
        )
    return jobs


def html_to_text(html_fragment):
    if not html_fragment:
        return ""
    text = unescape(html_fragment)
    text = re.sub(r"</(p|li|h[1-6]|div)>", "\n", text, flags=re.I)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"<li>", "• ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    lines = [ln.strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]
    return "\n".join(lines)


def parse_detail_page(html):
    """Pull the JobPosting JSON-LD block out of a job detail page."""
    match = re.search(
        r'<script type="application/ld\+json">\s*(\{.*?"@type":"JobPosting".*?\})\s*</script>',
        html,
        flags=re.S,
    )
    if not match:
        return {}
    try:
        data = json.loads(match.group(1))
    except json.JSONDecodeError:
        return {}

    company = ""
    hiring_org = data.get("hiringOrganization") or {}
    if isinstance(hiring_org, dict):
        company = hiring_org.get("name", "") or ""
    if not company or company.lower() in ("[hidden company]", "[unlock with premium]"):
        identifier = data.get("identifier") or {}
        if isinstance(identifier, dict):
            ident_name = identifier.get("name", "") or ""
            if ident_name and ident_name.lower() not in (
                "[hidden company]",
                "[unlock with premium]",
            ):
                company = ident_name

    return {
        "company": company,
        "description": html_to_text(data.get("description", "")),
        "date_posted": (data.get("datePosted") or "")[:10],
        "application_deadline": (data.get("validThrough") or "")[:10],
    }


def load_checkpoint():
    done = {}
    if os.path.exists(CHECKPOINT_CSV):
        with open(CHECKPOINT_CSV, newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                done[row["job_id"]] = row
    return done


def append_checkpoint(row, write_header):
    with open(CHECKPOINT_CSV, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        if write_header:
            writer.writeheader()
        writer.writerow(row)


def collect_job_list(target_count, search_term, max_pages=60):
    jobs = []
    seen_ids = set()
    page = 1
    while len(jobs) < target_count and page <= max_pages:
        url = f"{BASE}/remote-jobs?search={search_term}&page={page}"
        print(f"[listing] page {page} -> {url}")
        html = fetch(url)
        if html is None:
            print(f"  [error] could not fetch page {page}, stopping listing crawl")
            break
        page_jobs = parse_listing_page(html)
        if not page_jobs:
            print("  [info] no more job cards found, stopping listing crawl")
            break
        new_count = 0
        for job in page_jobs:
            if job["job_id"] not in seen_ids:
                seen_ids.add(job["job_id"])
                jobs.append(job)
                new_count += 1
        print(f"  [info] +{new_count} new jobs (total {len(jobs)})")
        page += 1
        time.sleep(DELAY)
    return jobs[:target_count]


def enrich_with_detail_pages(jobs):
    done = load_checkpoint()
    write_header = not os.path.exists(CHECKPOINT_CSV) or os.path.getsize(CHECKPOINT_CSV) == 0
    results = []
    total = len(jobs)
    for i, job in enumerate(jobs, start=1):
        if job["job_id"] in done:
            results.append(done[job["job_id"]])
            continue
        print(f"[detail {i}/{total}] {job['title'][:60]!r} ({job['url']})")
        html = fetch(job["url"])
        if html:
            detail = parse_detail_page(html)
            job.update({k: v for k, v in detail.items() if v})
        else:
            print(f"  [warn] failed to fetch detail page for job {job['job_id']}")
        append_checkpoint(job, write_header)
        write_header = False
        results.append(job)
        time.sleep(DELAY)
    return results


def export_to_excel(jobs, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [
        "Title",
        "Company",
        "Location",
        "Job Type",
        "Salary",
        "Experience Level",
        "Category",
        "Role Tag",
        "Skill Tags",
        "Date Posted",
        "Posted",
        "Application Deadline",
        "Summary",
        "Full Description",
        "Job URL",
    ]
    field_order = [
        "title",
        "company",
        "location",
        "job_type",
        "salary",
        "experience_level",
        "category",
        "role_tag",
        "skill_tags",
        "date_posted",
        "posted_relative",
        "application_deadline",
        "summary",
        "description",
        "url",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "DailyRemote Jobs"

    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for job in jobs:
        row = [job.get(f, "") for f in field_order]
        ws.append(row)

    url_col = field_order.index("url") + 1
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=url_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    # Descriptions run thousands of characters long; wrapping them blows every
    # row out to dozens of lines. Keep rows a uniform, compact single-line
    # height instead -- the full text is still in the cell (visible in the
    # formula bar or by widening the column), it just doesn't force the row
    # to expand.
    ws.sheet_format.defaultRowHeight = 18
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18
    for c in range(1, len(headers) + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=False, vertical="center"
            )

    widths = {
        "Title": 38,
        "Company": 22,
        "Location": 26,
        "Job Type": 14,
        "Salary": 22,
        "Experience Level": 16,
        "Category": 16,
        "Role Tag": 20,
        "Skill Tags": 30,
        "Date Posted": 13,
        "Posted": 14,
        "Application Deadline": 16,
        "Summary": 50,
        "Full Description": 60,
        "Job URL": 45,
    }
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 20)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=int, default=622, help="number of jobs to scrape")
    parser.add_argument("--search", default="", help="dailyremote search query")
    parser.add_argument(
        "--out", default=os.path.join(SCRIPT_DIR, "dailyremote_jobs.xlsx")
    )
    args = parser.parse_args()

    print(f"== Collecting up to {args.target} job listings ==")
    jobs = collect_job_list(args.target, args.search)
    print(f"== Found {len(jobs)} unique job listings, fetching detail pages ==")

    jobs = enrich_with_detail_pages(jobs)

    print(f"== Writing Excel workbook to {args.out} ==")
    export_to_excel(jobs, args.out)
    print("== Done ==")


if __name__ == "__main__":
    main()
